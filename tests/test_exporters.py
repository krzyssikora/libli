# tests/test_exporters.py
import csv
import datetime
import io
from decimal import Decimal

import openpyxl

from courses.exporters import _sanitize_text_cell
from courses.exporters import build_filename
from courses.exporters import to_csv
from courses.exporters import to_xlsx


def _matrix_table():
    return {
        "title": "Algebra — All my students — Progress",
        "subtitle": "Generated 2026-07-03 · Scope: All my students",
        "columns": [{"label": "Chapter 1", "max": None, "kind": "percent"}],
        "total_kind": "percent",
        "total_label": "Overall",
        "meta_row": None,
        "rows": [{"name": "Ada", "username": "ada", "cells": [85], "total": 85}],
        "footer": [{"label": "Average", "values": [85], "total": 85}],
    }


def _quiz_table():
    return {
        "title": "Algebra — Quiz gradebook",
        "subtitle": "Generated 2026-07-03 · Scope: All my students",
        "columns": [{"label": "1. Quiz", "max": Decimal("10"), "kind": "score"}],
        "total_kind": "score",
        "total_label": "Total",
        "meta_row": {"label": "Max", "values": [Decimal("10")], "total": Decimal("10")},
        "rows": [
            {
                "name": "=cmd()",
                "username": "ada",
                "cells": [Decimal("7")],
                "total": Decimal("7"),
            }
        ],
        "footer": [
            {"label": "Average", "values": [Decimal("7")], "total": Decimal("7")}
        ],
    }


def _read_csv(resp):
    body = b"".join(resp).decode("utf-8-sig")  # strips BOM
    return list(csv.reader(io.StringIO(body)))


def test_sanitize_neutralises_formula_prefixes():
    assert _sanitize_text_cell("=cmd()") == "'=cmd()"
    assert _sanitize_text_cell("+1") == "'+1"
    assert _sanitize_text_cell("-1") == "'-1"
    assert _sanitize_text_cell("@x") == "'@x"
    assert _sanitize_text_cell("Ada") == "Ada"
    assert _sanitize_text_cell(None) == ""


def test_build_filename():
    d = datetime.date(2026, 7, 3)
    assert (
        build_filename("algebra-i", "matrix", "results", False, d, "csv")
        == "algebra-i-matrix-results-2026-07-03.csv"
    )
    assert (
        build_filename("algebra-i", "quiz", "progress", False, d, "xlsx")
        == "algebra-i-quiz-2026-07-03.xlsx"
    )
    assert (
        build_filename("algebra-i", "quiz", "progress", True, d, "xlsx")
        == "algebra-i-quiz-numbers-2026-07-03.xlsx"
    )


def test_to_csv_matrix_percent_and_headers():
    resp = to_csv(_matrix_table(), "x.csv")
    assert resp["Content-Type"].startswith("text/csv")
    assert 'attachment; filename="x.csv"' in resp["Content-Disposition"]
    assert resp.content.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM
    rows = _read_csv(resp)
    assert rows[0] == ["Algebra — All my students — Progress"]
    assert rows[1] == ["Generated 2026-07-03 · Scope: All my students"]
    assert rows[3] == ["Name", "Username", "Chapter 1", "Overall"]
    assert rows[4] == ["Ada", "ada", "85%", "85%"]
    assert rows[5] == ["Average", "", "85%", "85%"]


def test_to_csv_quiz_scores_and_injection_guard():
    rows = _read_csv(to_csv(_quiz_table(), "q.csv"))
    assert rows[3] == ["Name", "Username", "1. Quiz", "Total"]
    assert rows[4] == ["Max", "", "10", "10"]  # meta Max row
    assert rows[5] == ["'=cmd()", "ada", "7", "7"]  # name neutralised, score numeric
    assert rows[6] == ["Average", "", "7", "7"]


def _load_xlsx(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.content))


def test_to_xlsx_scores_are_numeric_and_headers_present():
    resp = to_xlsx(_quiz_table(), "q.xlsx")
    assert resp["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'attachment; filename="q.xlsx"' in resp["Content-Disposition"]
    ws = _load_xlsx(resp).active
    # find the score cell for the student row (value 7 as a real number, not text)
    values = [c.value for col in ws.iter_cols() for c in col]
    assert 7 in values or 7.0 in values
    # injection guard applied to the =cmd() name
    assert any(isinstance(v, str) and v.startswith("'=cmd()") for v in values)


def test_to_xlsx_percent_cells_have_percent_format():
    resp = to_xlsx(_matrix_table(), "m.xlsx")
    ws = _load_xlsx(resp).active
    pct_cells = [c for col in ws.iter_cols() for c in col if c.number_format == "0%"]
    assert pct_cells and any(abs((c.value or 0) - 0.85) < 1e-9 for c in pct_cells)
