import csv

from django.http import HttpResponse
from django.utils.text import slugify
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_text_cell(value):
    """Neutralise CSV/XLSX formula injection: prefix a ' when a text token starts
    with a formula trigger. Numeric cells never pass through here."""
    text = "" if value is None else str(value)
    if text and text[0] in _DANGEROUS_PREFIXES:
        return "'" + text
    return text


def build_filename(slug, shape, mode, numbers_only, today, ext):
    """{slug}-{shape}-{mode? matrix only}-{numbers? quiz+numbers_only}-{ISO date}.{ext}.

    Filename shape: matrix mode is only appended for "matrix" shape; the
    "numbers" marker is only appended for "quiz" shape with numbers_only set.
    """
    parts = [slugify(slug) or "export", shape]
    if shape == "matrix":
        parts.append(mode)
    if shape == "quiz" and numbers_only:
        parts.append("numbers")
    parts.append(today.isoformat())
    return f"{'-'.join(parts)}.{ext}"


def _fmt_cell(value, kind):
    """Format a data/summary value by column kind for text output (CSV/HTML)."""
    if value is None:
        return ""
    if kind == "percent":
        return f"{value}%"
    if isinstance(value, str):  # a marker (—/…/R) — our own constant, safe
        return value
    return str(value)  # Decimal / int score


def to_csv(table, filename):
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(chr(0xFEFF))  # UTF-8 BOM for Excel (unambiguous — no invisible glyph)
    writer = csv.writer(resp)
    cols = table["columns"]
    tk = table["total_kind"]

    writer.writerow([_sanitize_text_cell(table["title"])])
    writer.writerow([_sanitize_text_cell(table["subtitle"])])
    writer.writerow([])
    writer.writerow(
        [_("Name"), _("Username")]
        + [_sanitize_text_cell(c["label"]) for c in cols]
        + [_sanitize_text_cell(table["total_label"])]
    )
    meta = table["meta_row"]
    if meta:
        writer.writerow(
            [
                _sanitize_text_cell(meta["label"]),
                "",
            ]  # label in Name col, Username blank
            + [
                _fmt_cell(v, c["kind"])
                for v, c in zip(meta["values"], cols, strict=True)
            ]
            + [_fmt_cell(meta["total"], tk)]
        )
    for row in table["rows"]:
        writer.writerow(
            [_sanitize_text_cell(row["name"]), _sanitize_text_cell(row["username"])]
            + [_fmt_cell(v, c["kind"]) for v, c in zip(row["cells"], cols, strict=True)]
            + [_fmt_cell(row["total"], tk)]
        )
    for frow in table["footer"]:
        writer.writerow(
            [_sanitize_text_cell(frow["label"]), ""]
            + [
                _fmt_cell(v, c["kind"])
                for v, c in zip(frow["values"], cols, strict=True)
            ]
            + [_fmt_cell(frow["total"], tk)]
        )
    return resp


def _xlsx_value(cell_value, kind):
    """Return (value, number_format) for an XLSX cell. Score -> float; percent ->
    fraction with a 0% format; marker/None -> sanitised text/empty."""
    if cell_value is None:
        return "", None
    if kind == "percent":
        return float(cell_value) / 100.0, "0%"
    if isinstance(cell_value, str):  # marker — safe constant
        return cell_value, None
    return float(cell_value), None  # Decimal/int score -> real number


def to_xlsx(table, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"
    cols = table["columns"]
    tk = table["total_kind"]
    bold = Font(bold=True)

    ws.append([_sanitize_text_cell(table["title"])])
    ws.append([_sanitize_text_cell(table["subtitle"])])
    ws.append([])

    header = (
        [_("Name"), _("Username")]
        + [_sanitize_text_cell(c["label"]) for c in cols]
        + [_sanitize_text_cell(table["total_label"])]
    )
    ws.append(header)
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.font = bold

    def _write_data_row(label_cells, values, total_value):
        ws.append(label_cells + [""] * (len(cols) + 1))  # placeholder, fill typed
        r = ws.max_row
        for j, (v, c) in enumerate(
            zip(values, cols, strict=True), start=len(label_cells) + 1
        ):
            value, fmt = _xlsx_value(v, c["kind"])
            ws.cell(row=r, column=j, value=value)
            if fmt:
                ws.cell(row=r, column=j).number_format = fmt
        value, fmt = _xlsx_value(total_value, tk)
        tcell = ws.cell(row=r, column=len(cols) + 3, value=value)
        if fmt:
            tcell.number_format = fmt
        return r

    meta = table["meta_row"]
    if meta:
        r = _write_data_row(
            [_sanitize_text_cell(meta["label"]), ""], meta["values"], meta["total"]
        )
        for cell in ws[r]:
            cell.font = bold
    for row in table["rows"]:
        _write_data_row(
            [_sanitize_text_cell(row["name"]), _sanitize_text_cell(row["username"])],
            row["cells"],
            row["total"],
        )
    for frow in table["footer"]:
        r = _write_data_row(
            [_sanitize_text_cell(frow["label"]), ""], frow["values"], frow["total"]
        )
        for cell in ws[r]:
            cell.font = Font(italic=True, bold=True)

    ws.freeze_panes = ws.cell(
        row=header_row_idx + 1, column=3
    )  # below header, right of identity cols

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp
